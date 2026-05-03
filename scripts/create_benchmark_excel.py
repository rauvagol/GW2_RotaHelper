import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


def load_benchmark_data(bench_dir: Path) -> list[dict[str, Any]]:
    benchmark_data = []

    subdirs = [
        Path("dps") / "power",
        Path("dps") / "condition",
        Path("alac") / "power",
        Path("alac") / "condition",
        Path("quick") / "power",
        Path("quick") / "condition",
    ]

    for subdir in subdirs:
        subdir_path = bench_dir / subdir
        if not subdir_path.exists():
            continue

        for json_file in subdir_path.glob("*.json"):
            try:
                with json_file.open("r", encoding="utf-8") as f:
                    data = json.load(f)

                if "buildMetadata" in data:
                    metadata = data["buildMetadata"].copy()
                    metadata["file_path"] = str(json_file.relative_to(bench_dir))
                    metadata["category"] = "_".join(subdir.parts)
                    if data.get("rotation"):
                        rotation = data["rotation"][0] if isinstance(data["rotation"][0][0], list) else data["rotation"]
                        metadata["rotation_length"] = len(rotation)
                        metadata["total_duration"] = rotation[-1][0] if rotation else 0
                    else:
                        metadata["rotation_length"] = 0
                        metadata["total_duration"] = 0
                    if "skillMap" in data:
                        metadata["unique_skills"] = len(data["skillMap"])
                    else:
                        metadata["unique_skills"] = 0

                    benchmark_data.append(metadata)

            except (json.JSONDecodeError, KeyError, FileNotFoundError) as e:
                print(f"Error processing {json_file}: {e}")
                continue

    return benchmark_data


def create_excel_report(benchmark_data: list[dict[str, Any]], output_file: Path) -> None:
    df = pd.DataFrame(benchmark_data)

    if df.empty:
        print("No benchmark data found!")
        return

    required_columns = {
        "name": "Unknown Build",
        "profession": "Unknown",
        "elite_spec": "Unknown",
        "build_type": "Unknown",
        "benchmark_type": "Unknown",
        "overall_dps": 0,
    }

    for col, default_val in required_columns.items():
        if col not in df.columns:
            df[col] = default_val
        else:
            df[col] = df[col].fillna(default_val)

    df["own_dps"] = 0
    df["bench_dps"] = df["overall_dps"]
    df["bench_percent"] = 0.0
    df["abs_95"] = 0
    df["rel_95"] = 0

    df = df.sort_values(["benchmark_type", "overall_dps"], ascending=[True, False])

    profession_colors = {
        "guardian": "FFB3D9FF",
        "warrior": "FFFFFF99",
        "engineer": "FFFF9933",
        "ranger": "FF99FF99",
        "thief": "FFB366FF",
        "elementalist": "FFFF6666",
        "mesmer": "FFCC66FF",
        "necromancer": "FF33CC33",
        "revenant": "FFFF9966",
    }

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        spec_to_profession = (
            df[["elite_spec", "profession"]]
            .drop_duplicates()
            .sort_values("profession")
            .reset_index(drop=True)
        )
        spec_to_profession.columns = ["Elite Spec", "Profession"]

        colors_data = {
            "Profession": [prof.title() for prof in profession_colors.keys()],
            "Color_Code": [color for color in profession_colors.values()],
            "Example": [prof.title() for prof in profession_colors.keys()],
        }
        colors_df = pd.DataFrame(colors_data)
        colors_df.to_excel(writer, sheet_name="Colors", index=False)

        colors_sheet = writer.sheets["Colors"]
        colors_header_fill = PatternFill(start_color="FFB3D1FF", end_color="FFB3D1FF", fill_type="solid")
        for col in range(1, 4):
            colors_sheet.cell(row=1, column=col).fill = colors_header_fill

        for idx, (prof, color) in enumerate(profession_colors.items(), 2):
            colors_sheet.cell(row=idx, column=3).fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )

        colors_sheet.cell(row=1, column=5).value = "Elite Spec"
        colors_sheet.cell(row=1, column=6).value = "Profession"
        colors_sheet.cell(row=1, column=5).fill = colors_header_fill
        colors_sheet.cell(row=1, column=6).fill = colors_header_fill
        for sidx, srow in spec_to_profession.iterrows():
            colors_sheet.cell(row=sidx + 2, column=5).value = srow["Elite Spec"].title()
            colors_sheet.cell(row=sidx + 2, column=6).value = srow["Profession"].title()
            prof_color = profession_colors.get(srow["Profession"].lower(), "FFFFFFFF")
            colors_sheet.cell(row=sidx + 2, column=6).fill = PatternFill(
                start_color=prof_color, end_color=prof_color, fill_type="solid"
            )

        benchmark_order = ["dps", "alac", "quick"]
        for bench_type in benchmark_order:
            if bench_type not in df["benchmark_type"].values:  # noqa: PD011
                continue

            type_df = df[df["benchmark_type"] == bench_type].copy()
            type_sheet_df = type_df[[
                "name",
                "build_type",
                "own_dps",
                "bench_dps",
                "bench_percent",
                "abs_95",
                "rel_95",
                "profession",
            ]].copy()

            type_sheet_df.columns = [
                "Build Name",
                "Build Type",
                "Own DPS",
                "Bench DPS",
                "Bench %",
                "95% Abs",
                "95% Rel",
                "Profession",
            ]

            sheet_name = f"{bench_type.upper()} Builds"
            type_sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            type_worksheet = writer.sheets[sheet_name]

            header_fill = PatternFill(start_color="FFB3D1FF", end_color="FFB3D1FF", fill_type="solid")
            for col in range(1, 9):
                type_worksheet.cell(row=1, column=col).fill = header_fill

            type_worksheet.auto_filter.ref = f"A1:G{len(type_df) + 1}"
            type_worksheet.column_dimensions["H"].hidden = True

            for prof, color in profession_colors.items():
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                rule = FormulaRule(formula=[f'LOWER($H2)="{prof}"'], fill=fill)
                type_worksheet.conditional_formatting.add(f"A2:A{len(type_df) + 1}", rule)

            for idx, row in type_df.iterrows():
                row_num = type_df.index.get_loc(idx) + 2
                type_worksheet.cell(row=row_num, column=5).value = f"=ROUND(C{row_num}/D{row_num}*100,2)"
                type_worksheet.cell(row=row_num, column=6).value = f"=ROUND(D{row_num}*0.95,0)"
                type_worksheet.cell(row=row_num, column=7).value = f"=ROUND(F{row_num}-C{row_num},0)"

    print(f"Excel report created: {output_file}")
    print(f"Total builds processed: {len(df)}")
    print(f"Benchmark types: {', '.join(df['benchmark_type'].unique())}")
    print(f"Professions: {', '.join(sorted(df['profession'].unique()))}")


def main() -> None:
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    bench_dir = project_root / "data" / "bench"
    output_dir = project_root / bench_dir
    output_file = output_dir / "benchmark_report.xlsx"

    output_dir.mkdir(parents=True, exist_ok=True)
    if not bench_dir.exists():
        print(f"Error: Benchmark directory not found: {bench_dir}")
        sys.exit(1)

    print(f"Loading benchmark data from: {bench_dir}")

    benchmark_data = load_benchmark_data(bench_dir)

    if not benchmark_data:
        print("No benchmark data found!")
        sys.exit(1)

    print(f"Found {len(benchmark_data)} benchmark files")

    create_excel_report(benchmark_data, output_file)

    print(f"\nReport saved to: {output_file}")


if __name__ == "__main__":
    main()
